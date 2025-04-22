# 1_ocr_only.py

# すべての画像に対してOCRで文字認識し、情報だけを得る
# バウンディングボックスと認識結果の文字は描画はしない

# 入力フォルダ ◯◯_png
# 出力フォルダ ◯◯_ocr_only

import os
import function_fin
import function_ocr

# ベースフォルダのパス
base_folder = "/Users/ikejimasouta/Intern/Bill_Analysis/"


# OCRで文字認識のみ行い、結果は描画しない
def process_png_folder(base_folder):
    for root, dirs, files in os.walk(base_folder):
        dirs = function_fin.sort_naturally(dirs)
        for dir_name in dirs:
            new_base_folder = os.path.join(root, dir_name)
            for root_2, dirs_2, files_2 in os.walk(new_base_folder):
                dirs_2 = function_fin.sort_naturally(dirs_2)
                files_2 = function_fin.sort_naturally(files_2)
                for dir_name_2 in dirs_2:
                    # フォルダ名が "UVDoc_png" で終わるかどうかを確認
                    # if dir_name_2.endswith("UVDoc_png"):
                    if dir_name_2.endswith("png") and "UVDoc" not in dir_name_2:
                        png_folder = os.path.join(root_2, dir_name_2)
                        input_folder = png_folder + "/"
                        # output = os.path.join(root_2, f"{dir_name_2[:8]}_ocr_only")
                        output = os.path.join(root_2, f"{dir_name_2[:2]}_ocr_only")
                        output_folder = output + "/"
                        if not os.path.exists(output_folder):
                            os.makedirs(output_folder)

                        # OCR処理を実行
                        function_ocr.ocr_only_function(input_folder, output_folder)
                break
        break


# 処理の実行
process_png_folder(base_folder)

# 2_change_direction.py

# 横向きの画像を回転させて方向を統一
# OCRの項目名情報をもとに横向きの画像を +90°,-90° 回転を選択

# 入力フォルダ ◯◯_ocr_only
# 出力フォルダ ◯◯_png_rot


import os
import re
import json
import numpy as np
import function_fin
import function_2_direction

# ベースフォルダのパス
base_folder = "/Users/ikejimasouta/Intern/Bill_Analysis/"

# キーワード群
words_1_a = ["単価", "数量", "診察項目名", "診療明細内容", "日付", "診療明細書"]
words_1_b = [
    "単価",
    "数量",
    "単",
    "数",
    "診療明細",
    "日",
    "ちゃん",
    "計算",
    "電話",
    "取引",
]
words_2_a = ["小計", "合計", "消費税", "消費", "外税"]
words_2_b = ["小計", "税", "計", "請求"]


def natural_sort_key(s):
    """自然順序でソートするためのキー関数"""
    return [
        int(text) if text.isdigit() else text.lower() for text in re.split(r"(\d+)", s)
    ]


def process_png_folder(base_folder):
    for root, dirs, files in os.walk(base_folder):
        dirs.sort(key=natural_sort_key)
        for dir_name in dirs:
            #  "数字UVDoc__ocr_only" で終わるディレクトリのみを処理
            # if dir_name.endswith("UVDoc_ocr_only"):
            if dir_name.endswith("ocr_only") and "UVDoc" not in dir_name:
                dir_path = os.path.join(root, dir_name)
                # output_folder = os.path.join(root, f"{dir_name[:8]}_png_rot")
                output_folder = os.path.join(root, f"{dir_name[:2]}_png_rot")
                os.makedirs(output_folder, exist_ok=True)

                for file_name in sorted(os.listdir(dir_path), key=natural_sort_key):
                    # jsonファイルに対して処理
                    if file_name.endswith(".json"):
                        file_path = os.path.join(dir_path, file_name)
                        print("-----------------------------------------")
                        with open(file_path) as f:
                            data = json.load(f)

                        # 縦横の比率を求める
                        prop = function_2_direction.calculate_width_prop(data)

                        # 各データを得る
                        (
                            text_list,
                            x_mean_list,
                            y_mean_list,
                            x_max_list,
                            y_max_list,
                            x_min_list,
                            y_min_list,
                        ) = function_fin.process_data(data)

                        print(file_name)
                        print(f"比率: {prop}")

                        x_max = np.max(x_max_list)
                        # キーワードの座標を取得しどちらに回転させるかを判定
                        x_1 = function_fin.find_word_y(
                            words_1_a, words_1_b, text_list, x_max_list, 0, x_max
                        )
                        x_2 = function_fin.find_word_y(
                            words_2_a, words_2_b, text_list, x_min_list, 0, x_max
                        )

                        # JSONファイルから画像ファイル名を取得
                        image_file_name = file_name.replace(
                            ".json", ".png"
                        )  # 拡張子を適切に変更
                        image_path = os.path.join(dir_path, image_file_name)

                        function_2_direction.change_direction(
                            prop, x_1, x_2, image_path, output_folder
                        )


# フォルダ内のすべてのPNGフォルダを処理
process_png_folder(base_folder)

# 3_make_flat.py

# すべての画像に対して、aspose.ocrを用いて画像を水平にする処理を実装
# 2_change_direction.pyで回転させた画像を水平にさせる

# 入力フォルダ ◯◯_png_rot
# 出力フォルダ ◯◯_png_flat

import os
import re
import aspose.ocr as ocr
from PIL import Image

# OCR APIのインスタンスを作成・ベースフォルダのパス作成
api = ocr.AsposeOcr()
base_folder = "/Users/ikejimasouta/Intern/Bill_Analysis/"


def natural_sort_key(s):
    """自然順序でソートするためのキー関数"""
    return [
        int(text) if text.isdigit() else text.lower() for text in re.split(r"(\d+)", s)
    ]


# aspose.ocrを用いて画像を水平にする
def process_png_folder(base_folder):
    for root, dirs, files in os.walk(base_folder):
        dirs.sort(key=natural_sort_key)
        for dir_name in dirs:
            # if dir_name.endswith("UVDoc_png_rot"):
            if dir_name.endswith("png_rot") and "UVDoc" not in dir_name:
                dir_path = os.path.join(root, dir_name)
                # output_folder = os.path.join(root, f"{dir_name[:8]}_png_flat")
                output_folder = os.path.join(root, f"{dir_name[:2]}_png_flat")
                for file_name in sorted(os.listdir(dir_path), key=natural_sort_key):
                    # ファイル名が ".png" で終わるかどうかを確認
                    if file_name.endswith(".png"):
                        image_path = os.path.join(dir_path, file_name)
                        if not os.path.exists(output_folder):
                            os.makedirs(output_folder)

                        # 画像入力のインスタンスを作成
                        img = ocr.OcrInput(ocr.InputType.SINGLE_IMAGE)
                        img.add(image_path)

                        # スキュー角度の検出
                        angles = api.calculate_skew(img)
                        for angle in angles:
                            print("-----------------------------------------")
                            print(f"File name: {file_name}")
                            print(f"Skew angle: {angle.angle:.2f}°")
                            skew_angle = angle.angle

                            # 画像を回転させるためにPillowライブラリを使用
                            original_image = Image.open(image_path)
                            rotated_image = original_image.rotate(
                                skew_angle, expand=True
                            )

                            # 処理した画像を保存
                            corrected_image_path = os.path.join(
                                output_folder,
                                f"{os.path.splitext(file_name)[0]}_flat.png",
                            )
                            rotated_image.save(corrected_image_path)


# 処理の実行
process_png_folder(base_folder)

# 4_ocr_all.py

# 平坦かつ水平になった画像に対してOCRで文字認識
# バウンディングボックス・認識結果も描画

# 入力フォルダ ◯◯_png_flat
# 出力フォルダ ◯◯_png_ocr

import os
import function_fin
import function_ocr


# ベースフォルダのパス
base_folder = "/Users/ikejimasouta/Intern/Bill_Analysis/"


# すべての画像に対してOCRで文字認識を行う
def process_png_folder(base_folder):
    for root, dirs, files in os.walk(base_folder):
        dirs = function_fin.sort_naturally(dirs)
        for dir_name in dirs:
            new_base_folder = os.path.join(root, dir_name)
            for root_2, dirs_2, files_2 in os.walk(new_base_folder):
                dirs_2 = function_fin.sort_naturally(dirs_2)
                files_2 = function_fin.sort_naturally(files_2)
                for dir_name_2 in dirs_2:
                    # if dir_name_2.endswith("UVDoc_png_flat"):
                    if dir_name_2.endswith("png_flat") and "UVDoc" not in dir_name_2:
                        png_folder = os.path.join(root_2, dir_name_2)
                        input_folder = png_folder + "/"
                        # output = os.path.join(root_2, f"{dir_name_2[:8]}_ocr")
                        output = os.path.join(root_2, f"{dir_name_2[:2]}_ocr")
                        output_folder = output + "/"
                        if not os.path.exists(output_folder):
                            os.makedirs(output_folder)
                        # OCR処理を実行
                        function_ocr.ocr_function(input_folder, output_folder)
                break
        break


# 処理の実行
process_png_folder(base_folder)

# 5_search_all.py

# すべてのOCR画像から"項目名・単価・個数・金額"の情報を取得する
# 独自アルゴリズムでマッチング
# 結果をターミナルに出力する

# 入力フォルダ ◯◯_png_ocr
# 出力：ターミナルに表示

import os
import function_fin
import function_5_search_one


# ベースフォルダのパス
base_folder = "/Users/ikejimasouta/Intern/Bill_Analysis/"


# OCRを利用して"項目名・単価・個数・金額"の情報を取得し、ターミナルに表示する
def process_png_folder(base_folder):
    for root, dirs, files in os.walk(base_folder):
        dirs = function_fin.sort_naturally(dirs)
        for dir_name in dirs:
            new_base_folder = os.path.join(root, dir_name)
            for root_2, dirs_2, files_2 in os.walk(new_base_folder):
                dirs_2 = function_fin.sort_naturally(dirs_2)
                files_2 = function_fin.sort_naturally(files_2)
                for dir_name_2 in dirs_2:
                    # if dir_name_2.endswith("UVDoc_ocr"):
                    if dir_name_2.endswith("_ocr") and "UVDoc" not in dir_name_2:

                        png_folder = os.path.join(root_2, dir_name_2)
                        input_folder = png_folder + "/"

                        print("           ")
                        print("           ")
                        print("           ")
                        print(
                            f"=={dir_name_2}================================================================================================"
                        )
                        function_5_search_one.search_all(input_folder)
                break
        break


# 処理を実行
process_png_folder(base_folder)

# 6_write_excel.py

# 明細部分+基本情報一部をexcelファイルに出力する
# アルゴリズムによる認識結果をresult_UVDoc.xlsxファイルに出力

# 入力フォルダ ◯◯_png_ocr
# 出力：エクセルファイルとして出力

import pandas as pd
import function_6_excel

# Excel ファイルに書き込む
# output_file_path = "result_ocr.xlsx"
output_file_path = "result_ocr_UVDoc.xlsx"

# ベースフォルダのパス
base_folder = "/Users/ikejimasouta/Intern/Bill_Analysis/"

# フォルダ内のすべてのOCRフォルダを処理
(
    ans_file,
    ans_item,
    ans_price,
    ans_num,
    ans_total,
    ans_date,
    ans_date_2,
    ans_number,
    ans_total_money,
    ans_tax,
) = function_6_excel.get_info(base_folder)

flat_file = [item for sublist in ans_file for item in sublist]
flat_item = [item for sublist in ans_item for item in sublist]
flat_price = [item for sublist in ans_price for item in sublist]
flat_num = [item for sublist in ans_num for item in sublist]
flat_total = [item for sublist in ans_total for item in sublist]
flat_date = [item for sublist in ans_date for item in sublist]
flat_date_2 = [item for sublist in ans_date_2 for item in sublist]
flat_number = [item for sublist in ans_number for item in sublist]
flat_total_money = [item for sublist in ans_total_money for item in sublist]
flat_tax = [item for sublist in ans_tax for item in sublist]


df = pd.DataFrame(
    columns=[
        "ファイル名",
        "項目",
        "単価",
        "個数",
        "金額",
        "発行日",
        "日付",
        "請求書No",
        "合計金額",
        "税金",
    ]
)

# 各リストの要素を1つずつDataFrameに追加
for i in range(len(flat_item)):
    new_row = pd.DataFrame(
        {
            "ファイル名": [flat_file[i]],
            "項目": [flat_item[i]],
            "単価": [flat_price[i]],
            "個数": [flat_num[i]],
            "金額": [flat_total[i]],
            "発行日": [flat_date[i]],
            "日付": [flat_date_2[i]],
            "請求書No": [flat_number[i]],
            "合計金額": [flat_total_money[i]],
            "税金": [flat_tax[i]],
        }
    )
    df = pd.concat([df, new_row], ignore_index=True)


df.to_excel(output_file_path, index=False, sheet_name="帳票")

# 7_document_intelligence.py

# !!注意!!
# 途中でエラーになることが多いがもう一度試すとできる(おそらく.DS_Storeが原因)
# それでもエラーになる場合は各フォルダ毎に Layout_all.py を実行する
# AzureのDocument Intelligenceを用いた認識
# 読取ることができる基本情報は、「小計」「合計」「総税額」「取得日」「取得時間」

# 入力フォルダ ◯◯_png
# 出力フォルダ ◯◯_png_xlsx

import os
import function_fin
import function_7_document_intelligence

# ベースフォルダのパス
base_folder = "/Users/ikejimasouta/Intern/Bill_Analysis/"


def process_png_folder(base_folder):
    for root, dirs, files in os.walk(base_folder):
        dirs = function_fin.sort_naturally(dirs)
        for dir_name in dirs:
            new_base_folder = os.path.join(root, dir_name)
            for root_2, dirs_2, files_2 in os.walk(new_base_folder):
                dirs_2 = function_fin.sort_naturally(dirs_2)
                files_2 = function_fin.sort_naturally(files_2)
                for dir_name_2 in dirs_2:
                    # フォルダ名が "png" で終わるかどうかを確認
                    # if dir_name_2.endswith("UVDoc_png"):
                    if dir_name_2.endswith("png") and "UVDoc" not in dir_name_2:
                        png_folder = os.path.join(root_2, dir_name_2)
                        input_folder = png_folder + "/"
                        # 表構造をエクセルで出力
                        function_7_document_intelligence.process_excel_files(
                            input_folder
                        )
                break
        break


# 処理を実行
process_png_folder(base_folder)

# 8_1_search_base_ocr.py

# エクセルファイルから基本情報の読取り
# OCRを利用し、自身のアルゴリズムでマッチング
# "~~~UVDoc_ocr_2"のファイル画像に対して探索

import os
import function_fin
import function_8_1_search_base_ocr

# ベースフォルダのパス
base_folder = "/Users/ikejimasouta/Intern/Bill_Analysis/"
output_folder = "/Users/ikejimasouta/Intern/Bill_Analysis/output_folder"
if not os.path.exists(output_folder):
    os.makedirs(output_folder)


# OCRを用いて基本情報を取得し、エクセルファイルに出力する
def process_png_folder(base_folder, output_folder):
    for root, dirs, files in os.walk(base_folder):
        dirs = function_fin.sort_naturally(dirs)
        for dir_name in dirs:
            new_base_folder = os.path.join(root, dir_name)
            for root_2, dirs_2, files_2 in os.walk(new_base_folder):
                dirs_2 = function_fin.sort_naturally(dirs_2)
                files_2 = function_fin.sort_naturally(files_2)
                for dir_name_2 in dirs_2:
                    if dir_name_2.endswith("UVDoc_ocr"):
                        # if dir_name_2.endswith("_ocr") and "UVDoc" not in dir_name_2:
                        png_folder = os.path.join(root_2, dir_name_2)
                        input_folder = png_folder + "/"
                        function_8_1_search_base_ocr.search_all_base(
                            input_folder, output_folder
                        )
                break
        break


# 処理を実行
process_png_folder(base_folder, output_folder)

# 8_2_search_base_doc_intelli.py

# Document Intelligenceを利用した基本情報の認識
# "~~~png_xlsx"フォルダ内の画像に対して処理
# 各フォルダ内のoutputフォルダに結果ファイルは格納される

# 入力フォルダ ◯◯_png_xlsx
# 出力フォルダ ◯◯/output/output_3_xlsx


import os
import function_fin
import function_8_2_search_base_doc_intelli

# ベースフォルダのパス
base_folder = "/Users/ikejimasouta/Intern/Bill_Analysis/"


# Document Intelligence を用いて基本情報を取得し、エクセルファイルに出力する
def process_png_folder(base_folder):
    for root, dirs, files in os.walk(base_folder):
        dirs = function_fin.sort_naturally(dirs)
        for dir_name in dirs:
            new_base_folder = os.path.join(root, dir_name)
            for root_2, dirs_2, files_2 in os.walk(new_base_folder):
                dirs_2 = function_fin.sort_naturally(dirs_2)
                files_2 = function_fin.sort_naturally(files_2)
                for dir_name_2 in dirs_2:
                    # フォルダ名が "_png_xlsx_3" で終わるかどうかを確認
                    # if dir_name_2.endswith("UVDoc_png_xlsx"):
                    if dir_name_2.endswith("_png_xlsx") and "UVDoc" not in dir_name_2:
                        png_folder = os.path.join(root_2, dir_name_2)
                        input_folder = png_folder + "/"
                        function_8_2_search_base_doc_intelli.extract_total_values_from_excel(
                            input_folder
                        )
                break
        break


# 処理を実行
process_png_folder(base_folder)

# 9_connect_base

# Document Intelligenceの認識結果をまとめるためのファイル
# フォルダごとの基本情報に関するエクセルファイルを一つのエクセルファイルに結合する
# outputフォルダ内の結果を一つのエクセルファイルにまとめる

import os
import function_fin
import function_9_connect

# ベースフォルダのパス
base_folder = "/Users/ikejimasouta/Intern/Bill_Analysis/"
output_folder = "/Users/ikejimasouta/Intern/Bill_Analysis/output_folder"
if not os.path.exists(output_folder):
    os.makedirs(output_folder)


# 11フォルダにそれぞれ存在するエクセルファイルを1つのエクセルファイルとして統合する
def process_png_folder(base_folder, output_folder):
    for root, dirs, files in os.walk(base_folder):
        dirs = function_fin.sort_naturally(dirs)
        for dir_name in dirs:
            new_base_folder = os.path.join(root, dir_name)
            for root_2, dirs_2, files_2 in os.walk(new_base_folder):
                dirs_2 = function_fin.sort_naturally(dirs_2)
                files_2 = function_fin.sort_naturally(files_2)
                for dir_name_2 in dirs_2:
                    if dir_name_2.endswith("output"):
                        png_folder = os.path.join(root_2, dir_name_2)
                        input_folder = png_folder + "/"
                        function_9_connect.connect(input_folder, output_folder)
                break
        break


# 処理を実行
process_png_folder(base_folder, output_folder)

# 10_1_merge_excel.py

# azure Document Intelligence, 自身が実装したアルゴリズムの結果を統合する
# 基本項目に関して4手法の統合


import openpyxl


# 優先度1:UVDocありDocument Intelligence
file1_path = "/Users/ikejimasouta/Intern/Bill_Analysis/output_folder/output_UVDoc_doc_intelli_2.xlsx"

# 優先度2:UVDocなしDocument Intelligence
file2_path = (
    "/Users/ikejimasouta/Intern/Bill_Analysis/output_folder/output_doc_intelli_2.xlsx"
)

# 優先度3:UVDocなしOCR
file3_path = (
    "/Users/ikejimasouta/Intern/Bill_Analysis/output_folder/output_UVDoc_ocr.xlsx"
)

# 優先度4:UVDocありOCR
file4_path = (
    "/Users/ikejimasouta/Intern/Bill_Analysis/output_folder/output_png_ocr.xlsx"
)

# 出力ファイル
output_path = "/Users/ikejimasouta/Intern/Bill_Analysis/output_folder/output_merged_4methods_2.xlsx"


def replace_missing_values(file1_path, file2_path, output_path):
    # ファイル1を読み込み
    wb1 = openpyxl.load_workbook(file1_path)
    ws1 = wb1.active

    # ファイル2を読み込み
    wb2 = openpyxl.load_workbook(file2_path)
    ws2 = wb2.active

    # ファイル3を読み込み
    wb3 = openpyxl.load_workbook(file3_path)
    ws3 = wb3.active

    # ファイル4を読み込み
    wb4 = openpyxl.load_workbook(file4_path)
    ws4 = wb4.active

    max_row = max(ws1.max_row, ws2.max_row)
    max_col = max(ws1.max_column, ws2.max_column)

    # C列(税額)の置き換え
    for row in range(1, max_row + 1):
        cell1 = ws1.cell(row=row, column=3)
        cell2 = ws2.cell(row=row, column=3)
        cell3 = ws3.cell(row=row, column=3)
        cell4 = ws4.cell(row=row, column=3)

        if cell1.value == "*":
            cell1.value = cell2.value

        if cell1.value == "*":
            cell1.value = cell3.value

        if cell1.value == "*":
            cell1.value = cell4.value

    # C列(小計)の置き換え
    for row in range(1, max_row + 1):
        cell1 = ws1.cell(row=row, column=2)
        cell2 = ws2.cell(row=row, column=2)
        cell3 = ws3.cell(row=row, column=2)
        cell4 = ws4.cell(row=row, column=2)

        if cell1.value == "*":
            cell1.value = cell2.value

        if cell1.value == "*":
            cell1.value = cell3.value

        if cell1.value == "*":
            cell1.value = cell4.value

    # D列(合計)の置き換え
    for row in range(1, max_row + 1):
        cell1 = ws1.cell(row=row, column=4)
        cell2 = ws2.cell(row=row, column=4)
        cell3 = ws3.cell(row=row, column=4)
        cell4 = ws4.cell(row=row, column=4)

        if cell1.value == "*":
            cell1.value = cell2.value

        if cell1.value == "*":
            cell1.value = cell3.value

        if cell1.value == "*":
            cell1.value = cell4.value

    for row in range(1, max_row + 1):
        cell1 = ws1.cell(row=row, column=6)
        cell2 = ws2.cell(row=row, column=6)
        cell3 = ws3.cell(row=row, column=6)
        cell4 = ws4.cell(row=row, column=6)

        if cell1.value == "*":
            cell1.value = cell2.value

        if cell1.value == "*":
            cell1.value = cell3.value

        if cell1.value == "*":
            cell1.value = cell4.value

    # 結果を新しいファイルに保存
    wb1.save(output_path)
    print(f"completed")


replace_missing_values(file1_path, file2_path, output_path)

# 10_2_merge_excel.py

# 税額の精度を向上させる
# 4手法を統合して得られた合計金額の情報をもとに、税額を探索する
# 基本項目の統合

import openpyxl


# 優先度1:UVDocありDocument Intelligence
file1_path = "/Users/ikejimasouta/Intern/Bill_Analysis/output_folder/output_UVDoc_doc_intelli_2.xlsx"

# 優先度2:UVDocなしDocument Intelligence
file2_path = (
    "/Users/ikejimasouta/Intern/Bill_Analysis/output_folder/output_doc_intelli_2.xlsx"
)

# 優先度3:UVDocなしOCR
file3_path = (
    "/Users/ikejimasouta/Intern/Bill_Analysis/output_folder/output_UVDoc_ocr.xlsx"
)

# 優先度4:UVDocありOCR
file4_path = (
    "/Users/ikejimasouta/Intern/Bill_Analysis/output_folder/output_png_ocr.xlsx"
)

# 4手法統合したファイル
integrated_path = (
    "/Users/ikejimasouta/Intern/Bill_Analysis/output_folder/output_merged_4methods.xlsx"
)


# 出力ファイル
output_path = "/Users/ikejimasouta/Intern/Bill_Analysis/output_folder/output_merged_4methods_fin.xlsx"


def replace_missing_values(
    file1_path, file2_path, file3_path, file4_path, integrated_path, output_path
):
    # ファイル1を読み込み
    wb1 = openpyxl.load_workbook(file1_path)
    ws1 = wb1.active

    # ファイル2を読み込み
    wb2 = openpyxl.load_workbook(file2_path)
    ws2 = wb2.active

    # ファイル3を読み込み
    wb3 = openpyxl.load_workbook(file3_path)
    ws3 = wb3.active

    # ファイル4を読み込み
    wb4 = openpyxl.load_workbook(file4_path)
    ws4 = wb4.active

    # ファイル4を読み込み
    wb_integrated = openpyxl.load_workbook(integrated_path)
    ws_integrated = wb_integrated.active

    max_row = max(ws1.max_row, ws2.max_row)
    max_col = max(ws1.max_column, ws2.max_column)

    # C列(税額)の置き換え
    for row in range(1, max_row + 1):
        cell1 = ws1.cell(row=row, column=3)
        cell2 = ws2.cell(row=row, column=3)
        cell3 = ws3.cell(row=row, column=3)
        cell4 = ws4.cell(row=row, column=3)
        cell_integrated = ws_integrated.cell(row=row, column=4)

        try:
            threshold = int(cell_integrated.value) / 2
        except ValueError:
            threshold = 10000000  # デフォルトのしきい値

        if cell1.value == "*" and cell2.value != "*" and int(cell2.value) < threshold:
            cell1.value = cell2.value

        if cell1.value == "*" and cell3.value != "*" and int(cell3.value) < threshold:
            cell1.value = cell3.value

        if cell1.value == "*" and cell4.value != "*" and int(cell4.value) < threshold:
            cell1.value = cell4.value

    # C列(小計)の置き換え
    for row in range(1, max_row + 1):
        cell1 = ws1.cell(row=row, column=2)
        cell2 = ws2.cell(row=row, column=2)
        cell3 = ws3.cell(row=row, column=2)
        cell4 = ws4.cell(row=row, column=2)

        if cell1.value == "*":
            cell1.value = cell2.value

        if cell1.value == "*":
            cell1.value = cell3.value

        if cell1.value == "*":
            cell1.value = cell4.value

    # D列(合計)の置き換え
    for row in range(1, max_row + 1):
        cell1 = ws1.cell(row=row, column=4)
        cell2 = ws2.cell(row=row, column=4)
        cell3 = ws3.cell(row=row, column=4)
        cell4 = ws4.cell(row=row, column=4)

        if cell1.value == "*":
            cell1.value = cell2.value

        if cell1.value == "*":
            cell1.value = cell3.value

        if cell1.value == "*":
            cell1.value = cell4.value

    for row in range(1, max_row + 1):
        cell1 = ws1.cell(row=row, column=6)
        cell2 = ws2.cell(row=row, column=6)
        cell3 = ws3.cell(row=row, column=6)
        cell4 = ws4.cell(row=row, column=6)

        if cell1.value == "*":
            cell1.value = cell2.value

        if cell1.value == "*":
            cell1.value = cell3.value

        if cell1.value == "*":
            cell1.value = cell4.value

    # 結果を新しいファイルに保存
    wb1.save(output_path)
    print(f"completed")


replace_missing_values(
    file1_path, file2_path, file3_path, file4_path, integrated_path, output_path
)

# 11_merge_excel.py

# 請求書番号と入金額を求める
# azure Document Intelligenceの結果を統合する
# UVDocを利用した場合と利用しない場合の結果を統合する

import openpyxl


# 例としての使用方法:
file1_path = (
    "/Users/ikejimasouta/Intern/Bill_Analysis/output_folder/output_png_ocr.xlsx"
)
file2_path = (
    "/Users/ikejimasouta/Intern/Bill_Analysis/output_folder/output_UVDoc_ocr.xlsx"
)
output_path = "/Users/ikejimasouta/Intern/Bill_Analysis/output_folder/output_ocr.xlsx"


def replace_missing_values(file1_path, file2_path, output_path):
    # ファイル1を読み込み
    wb1 = openpyxl.load_workbook(file1_path)
    ws1 = wb1.active

    # ファイル2を読み込み
    wb2 = openpyxl.load_workbook(file2_path)
    ws2 = wb2.active

    max_row = max(ws1.max_row, ws2.max_row)
    max_col = max(ws1.max_column, ws2.max_column)

    # 請求書番号の置き換え
    for row in range(1, max_row + 1):
        cell1 = ws1.cell(row=row, column=7)
        cell2 = ws2.cell(row=row, column=7)

        if cell1.value == "*":
            cell1.value = cell2.value

    # 入金額の置き換え
    for row in range(1, max_row + 1):
        cell1 = ws1.cell(row=row, column=8)
        cell2 = ws2.cell(row=row, column=8)

        if cell1.value == "*":
            cell1.value = cell2.value

    # 結果を新しいファイルに保存
    wb1.save(output_path)
    print(f"completed")


replace_missing_values(file1_path, file2_path, output_path)

# 基本的な表構造の読取り

import os

from azure.ai.documentintelligence import DocumentIntelligenceClient
import pandas as pd
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import function_fin


endpoint = "https://infodeliver-formai-us.cognitiveservices.azure.com/"
key = "d7521137542c408386de9180b5da18a9"





def apply_borders(sheet, start_row, start_col, end_row, end_col):
    """Apply borders to a specified range in an Excel sheet"""
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    thick_border = Border(
        left=Side(style="thick", color="000000"),
        right=Side(style="thick", color="000000"),
        top=Side(style="thick", color="000000"),
        bottom=Side(style="thick", color="000000"),
    )

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.border = thin_border

    for col in range(start_col, end_col + 1):
        top_cell = sheet.cell(row=start_row, column=col)
        bottom_cell = sheet.cell(row=end_row, column=col)
        top_cell.border = thick_border
        bottom_cell.border = thick_border

    for row in range(start_row, end_row + 1):
        left_cell = sheet.cell(row=row, column=start_col)
        right_cell = sheet.cell(row=row, column=end_col)
        left_cell.border = thick_border
        right_cell.border = thick_border


def adjust_column_widths(sheet):
    """Adjust the width of columns based on their content"""
    default_font = Font(name="Calibri", size=11)
    approx_char_width = 1.2

    for col in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                cell_length = len(str(cell.value))
                char_multiplier = (
                    1 if not cell.font or not cell.font.sz else cell.font.sz / 11
                )
                max_length = max(max_length, cell_length * char_multiplier)

        adjusted_width = (max_length * approx_char_width) + 2
        sheet.column_dimensions[column_letter].width = adjusted_width


def process_invoices(input_folder):
    output_folder = f"{os.path.dirname(input_folder)}_xlsx_3"
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    for filename in os.listdir(input_folder):
        file_path = os.path.join(input_folder, filename)
        if os.path.isfile(file_path) and filename != ".DS_Store":
            print(f"Processing file: {filename}")
            with open(file_path, "rb") as f:
                poller = client.begin_analyze_document(
                    "prebuilt-invoice",
                    analyze_request=f,
                    content_type="application/octet-stream",
                )
            result = poller.result()

            if not result.tables:
                print(f"No tables found in {filename}. Skipping...")
                empty_table = pd.DataFrame([["*" for _ in range(1)] for _ in range(1)])
                excel_filename = os.path.join(
                    output_folder, f"0{os.path.splitext(filename)[0]}.xlsx"
                )
                with pd.ExcelWriter(excel_filename, engine="openpyxl") as writer:
                    empty_table.to_excel(writer, index=False, header=False)
                    apply_borders(writer.book.active, 1, 1, 1, 1)
                    adjust_column_widths(writer.book.active)
                continue

            tables = []
            for table in result.tables:
                df = pd.DataFrame(
                    index=range(table.row_count), columns=range(table.column_count)
                )
                for cell in table.cells:
                    df.iat[cell.row_index, cell.column_index] = cell.content
                df.fillna("", inplace=True)
                tables.append(df)

            excel_filename = os.path.join(
                output_folder, f"0{os.path.splitext(filename)[0]}.xlsx"
            )

            with pd.ExcelWriter(excel_filename, engine="openpyxl") as writer:
                start_row = 1
                for df in tables:
                    start_col = 1
                    df.to_excel(
                        writer,
                        startrow=start_row - 1,
                        startcol=start_col - 1,
                        index=False,
                        header=False,
                    )
                    end_row = start_row + df.shape[0] - 1
                    end_col = start_col + df.shape[1] - 1
                    apply_borders(
                        writer.book.active, start_row, start_col, end_row, end_col
                    )
                    start_row = end_row + 2  # 次のテーブルの開始行

                adjust_column_widths(writer.book.active)

    print("All files have been processed.")


# ベースフォルダのパス
base_folder = "/Users/ikejimasouta/Intern/Bill_Analysis/"


def process_png_folder(base_folder):
    for root, dirs, files in os.walk(base_folder):
        dirs = function_fin.sort_naturally(dirs)
        for dir_name in dirs:
            new_base_folder = os.path.join(root, dir_name)
            for root_2, dirs_2, files_2 in os.walk(new_base_folder):
                dirs_2 = function_fin.sort_naturally(dirs_2)
                files_2 = function_fin.sort_naturally(files_2)
                for dir_name_2 in dirs_2:
                    # フォルダ名が "png" で終わるかどうかを確認
                    if dir_name_2.endswith("png") and "UVDoc" not in dir_name_2:
                        png_folder = os.path.join(root_2, dir_name_2)
                        input_folder = png_folder + "/"
                        # 表構造をエクセルで出力
                        process_invoices(input_folder)
                break
        break


# 処理を実行
process_png_folder(base_folder)